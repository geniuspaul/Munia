# users/mixins.py

import openpyxl
from django.http import HttpResponse

class ExcelExportMixin:
    excel_export_fields = []  # List of model fields or related fields (e.g. user__email)
    excel_export_headers = []  # Optional: user-friendly headers
    excel_export_filename = "exported_data"

    def export_selected_to_excel(self, request, queryset):
        field_names = self.excel_export_fields
        headers = self.excel_export_headers or field_names

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Write header
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(field_names, 1):
                value = obj
                for attr in field.split('__'):
                    value = getattr(value, attr, '')
                ws.cell(row=row_num, column=col_num).value = str(value) if value is not None else ''

        # Prepare HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.excel_export_filename}.xlsx"'
        wb.save(response)
        return response

    export_selected_to_excel.short_description = "Export selected records to Excel"
